import os
import easyocr
from googletrans import Translator
from transformers import BlipProcessor, BlipForConditionalGeneration
from PIL import Image
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import torch

# Check if CUDA is available
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
if not torch.cuda.is_available():
    print("Warning: CUDA is not available. Running on CPU.")
else:
    print(f"Using GPU: {torch.cuda.get_device_name(0)}")

analyzer = SentimentIntensityAnalyzer()


def extract_text_with_easyocr(image_path):
    reader = easyocr.Reader(["en", "tl"], gpu=True)  # Enable GPU
    result = reader.readtext(image_path)
    extracted_text = " ".join([item[1] for item in result])
    return extracted_text


def translate_text(text, target_language="en"):
    if not text.strip():
        return ""
    try:
        translator = Translator()
        translated = translator.translate(text, dest=target_language)
        return translated.text
    except Exception as e:
        print(f"Error during translation: {e}")
        return text


def generate_caption(image_path):
    processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
    model = BlipForConditionalGeneration.from_pretrained(
        "Salesforce/blip-image-captioning-base"
    ).to(
        device
    )  # Move model to GPU

    image = Image.open(image_path)
    inputs = processor(images=image, return_tensors="pt")
    # Move inputs to GPU
    inputs = {k: v.to(device) for k, v in inputs.items()}

    out = model.generate(**inputs)
    caption = processor.decode(out[0], skip_special_tokens=True)
    return caption


# Rest of the code remains the same
def analyze_sentiment(text):
    sentiment = analyzer.polarity_scores(text)
    compound_score = sentiment["compound"]
    if compound_score >= 0.05:
        sentiment_label = "Positive"
    elif compound_score <= -0.05:
        sentiment_label = "Negative"
    else:
        sentiment_label = "Neutral"
    return (
        sentiment_label,
        sentiment["neg"],
        sentiment["neu"],
        sentiment["pos"],
        compound_score,
    )


def calculate_confidence(overall_score):
    confidence = abs(overall_score) * 100
    if confidence < 50:
        confidence = 50 + (confidence * 0.5)
    else:
        confidence = 50 + (confidence**1.8) / 20
    confidence = min(confidence, 100)
    return round(confidence, 2)


def create_or_load_excel(output_folder, results):
    excel_file = os.path.join(output_folder, "sentiment_analysis_results.xlsx")
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sentiment Analysis Results"
        sheet.append(results.columns.tolist())
    for row in dataframe_to_rows(results, index=False, header=False):
        sheet.append(row)
    workbook.save(excel_file)
    print(f"Results saved to {excel_file}")


def process_folder(folder_path):
    results = []
    error_count = 0
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):
            try:
                extracted_text = extract_text_with_easyocr(file_path)
                if not extracted_text:
                    translated_text = ""
                    sentiment_translated = "Neutral"
                    score_translated = 0
                else:
                    translated_text = translate_text(extracted_text, "en")
                    (
                        sentiment_translated,
                        neg_translated,
                        neu_translated,
                        pos_translated,
                        score_translated,
                    ) = analyze_sentiment(translated_text)
                image_caption = generate_caption(file_path)
                (
                    sentiment_caption,
                    neg_caption,
                    neu_caption,
                    pos_caption,
                    score_caption,
                ) = analyze_sentiment(image_caption)
                overall_score = score_translated + score_caption
                if overall_score >= 0.05:
                    overall_sentiment = "Positive"
                elif overall_score <= -0.05:
                    overall_sentiment = "Negative"
                else:
                    overall_sentiment = "Neutral"
                confidence_level = calculate_confidence(overall_score)
                results.append(
                    {
                        "File Name": filename,
                        "Extracted Text": extracted_text,
                        "Translated Text": translated_text,
                        "Text Sentiment": sentiment_translated,
                        "Image Caption": image_caption,
                        "Caption Sentiment": sentiment_caption,
                        "Overall Sentiment": overall_sentiment,
                        "Confidence": f"{confidence_level}%",
                    }
                )
                print(f"Processed: {filename}")
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                error_count += 1
                continue
    results_df = pd.DataFrame(results)
    output_folder = os.path.dirname(folder_path)
    create_or_load_excel(output_folder, results_df)
    print(f"Processing complete. {error_count} image(s) were skipped due to errors.")
    sentiment_counts = {"Positive": 0, "Neutral": 0, "Negative": 0}
    for result in results:
        sentiment = result["Overall Sentiment"]
        if sentiment in sentiment_counts:
            sentiment_counts[sentiment] += 1
    labels = sentiment_counts.keys()
    counts = sentiment_counts.values()
    fig, ax = plt.subplots()
    bars = ax.bar(labels, counts, color=["green", "gray", "red"])
    ax.set_title("Sentiment Distribution")
    ax.set_xlabel("Sentiment")
    ax.set_ylabel("Count")
    for bar in bars:
        yval = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            yval + 0.1,
            int(yval),
            ha="center",
            va="bottom",
        )
    plt.savefig(os.path.join(output_folder, "sentiment_distribution.png"))
    plt.show()


if __name__ == "__main__":
    choice = (
        input(
            "Do you want to process a specific image or a folder? (Enter 'image' or 'folder'): "
        )
        .strip()
        .lower()
    )
    if choice == "folder":
        folder_path = input("Enter the folder path: ").strip()
        if os.path.isdir(folder_path):
            process_folder(folder_path)
        else:
            print(f"The folder '{folder_path}' does not exist.")
    elif choice == "image":
        image_path = input("Enter the image file path: ").strip()
        if os.path.isfile(image_path):
            print(f"Processing Image: {image_path}")
            extracted_text = extract_text_with_easyocr(image_path)
            print(f"Extracted Text: {extracted_text}")
            translated_text = translate_text(extracted_text, "en")
            print(f"Translated Text: {translated_text}")
            (
                sentiment_translated,
                neg_translated,
                neu_translated,
                pos_translated,
                score_translated,
            ) = analyze_sentiment(translated_text)
            print(
                f"Sentiment of Translated Text: {sentiment_translated} (neg={neg_translated}, neu={neu_translated}, pos={pos_translated}, Score: {score_translated})"
            )
            image_caption = generate_caption(image_path)
            print(f"Image Caption: {image_caption}")
            sentiment_caption, neg_caption, neu_caption, pos_caption, score_caption = (
                analyze_sentiment(image_caption)
            )
            print(
                f"Sentiment of Image Caption: {sentiment_caption} (neg={neg_caption}, neu={neu_caption}, pos={pos_caption}, Score: {score_caption})"
            )
            overall_score = score_translated + score_caption
            if overall_score >= 0.05:
                overall_sentiment = "Positive"
            elif overall_score <= -0.05:
                overall_sentiment = "Negative"
            else:
                overall_sentiment = "Neutral"
            print(f"Overall Sentiment: {overall_sentiment} (Score: {overall_score})")
            confidence_level = calculate_confidence(overall_score)
            print(f"Confidence Level: {confidence_level}%")
        else:
            print(f"The file '{image_path}' does not exist.")
    else:
        print("Invalid choice! Please enter 'image' or 'folder'.")
