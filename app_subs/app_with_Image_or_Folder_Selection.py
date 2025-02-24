import os
import easyocr
from googletrans import Translator
from transformers import BlipProcessor, BlipForConditionalGeneration
from PIL import Image
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import torch


# Initialize the VADER sentiment analyzer
analyzer = SentimentIntensityAnalyzer()


def extract_text_with_easyocr(image_path):
    """Extract text from an image using EasyOCR"""
    reader = easyocr.Reader(["en", "tl"], gpu=True)
    result = reader.readtext(image_path)
    # Extracting text from the result
    extracted_text = " ".join([item[1] for item in result])
    return extracted_text


def translate_text(text, target_language="en"):
    """Translates extracted text to the target language"""
    if not text.strip():  # Check if the text is empty or whitespace
        return ""  # Return empty string if no text to translate
    try:
        translator = Translator()
        translated = translator.translate(text, dest=target_language)
        return translated.text
    except Exception as e:
        print(f"Error during translation: {e}")
        return text  # Fallback to original text in case of an error


def generate_caption(image_path):
    """Generate a caption using BLIP"""
    processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
    model = BlipForConditionalGeneration.from_pretrained(
        "Salesforce/blip-image-captioning-base"
    )

    # Open the image
    image = Image.open(image_path)

    # Preprocess the image and generate the caption
    inputs = processor(images=image, return_tensors="pt")
    out = model.generate(**inputs)

    # Decode and return the caption
    caption = processor.decode(out[0], skip_special_tokens=True)
    return caption


def analyze_sentiment(text):
    """Analyze sentiment using VADER"""
    sentiment = analyzer.polarity_scores(text)
    compound_score = sentiment["compound"]

    if compound_score >= 0.05:
        sentiment_label = "Positive"
    elif compound_score <= -0.05:
        sentiment_label = "Negative"
    else:
        sentiment_label = "Neutral"

    return (
        sentiment_label,
        sentiment["neg"],
        sentiment["neu"],
        sentiment["pos"],
        compound_score,
    )


def calculate_confidence(overall_score):
    """Calculate confidence level as percentage"""
    confidence = abs(overall_score) * 100  # Multiply by 100 to get percentage

    # Apply exponential scaling to boost the confidence for stronger scores
    if confidence < 50:
        confidence = 50 + (
            confidence * 0.5
        )  # Gradually increase confidence for smaller scores
    else:
        confidence = (
            50 + (confidence**1.8) / 20
        )  # Exponential boost for stronger confidence

    # Cap the confidence to 100%
    confidence = min(confidence, 100)

    return round(
        confidence, 2
    )  # Return confidence as a percentage (round to 2 decimal places)


def create_or_load_excel(output_folder, results):
    """Create or load the Excel file and append new results"""
    excel_file = os.path.join(output_folder, "sentiment_analysis_results.xlsx")

    # If file exists, load it; otherwise, create it and add headers
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sentiment Analysis Results"
        sheet.append(results.columns.tolist())  # Add headers to the sheet

    # Append new results to the sheet
    for row in dataframe_to_rows(results, index=False, header=False):
        sheet.append(row)

    # Save the updated Excel file
    workbook.save(excel_file)
    print(f"Results saved to {excel_file}")


def process_folder(folder_path):
    """Process images in a folder and save results to Excel"""
    results = []
    error_count = 0

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)

        # Check if the file is an image
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):
            try:
                extracted_text = extract_text_with_easyocr(file_path)

                # If no text is extracted, assign neutral sentiment
                if not extracted_text:
                    translated_text = ""
                    sentiment_translated = "Neutral"
                    score_translated = 0
                else:
                    # Translate extracted text if it's in Tagalog (Filipino)
                    translated_text = translate_text(
                        extracted_text, "en"
                    )  # Translate to English
                    (
                        sentiment_translated,
                        neg_translated,
                        neu_translated,
                        pos_translated,
                        score_translated,
                    ) = analyze_sentiment(translated_text)

                # Generate a caption describing the image using BLIP
                image_caption = generate_caption(file_path)

                # Analyze sentiment of the image caption
                (
                    sentiment_caption,
                    neg_caption,
                    neu_caption,
                    pos_caption,
                    score_caption,
                ) = analyze_sentiment(image_caption)

                # Calculate overall sentiment based on both scores (translated text + image caption)
                overall_score = score_translated + score_caption
                if overall_score >= 0.05:
                    overall_sentiment = "Positive"
                elif overall_score <= -0.05:
                    overall_sentiment = "Negative"
                else:
                    overall_sentiment = "Neutral"

                # Calculate confidence level as percentage
                confidence_level = calculate_confidence(overall_score)

                # Add data to the results list
                results.append(
                    {
                        "File Name": filename,
                        "Extracted Text": extracted_text,
                        "Translated Text": translated_text,
                        "Text Sentiment": sentiment_translated,
                        "Image Caption": image_caption,
                        "Caption Sentiment": sentiment_caption,
                        "Overall Sentiment": overall_sentiment,
                        "Confidence": f"{confidence_level}%",  # Add % sign to confidence
                    }
                )

                print(f"Processed: {filename}")

            except Exception as e:
                print(f"Error processing {filename}: {e}")
                error_count += 1  # Increment error count
                continue  # Skip the corrupted or errored image and continue with the next one

    # Convert the results to a DataFrame
    results_df = pd.DataFrame(results)

    # Create or load the Excel file and append new results
    output_folder = os.path.dirname(folder_path)
    create_or_load_excel(output_folder, results_df)

    # Report errors
    print(f"Processing complete. {error_count} image(s) were skipped due to errors.")

    # Create a bar graph from the overall sentiments
    sentiment_counts = {"Positive": 0, "Neutral": 0, "Negative": 0}

    # Count occurrences of each sentiment
    for result in results:
        sentiment = result["Overall Sentiment"]
        if sentiment in sentiment_counts:
            sentiment_counts[sentiment] += 1

    # Plotting the bar graph
    labels = sentiment_counts.keys()
    counts = sentiment_counts.values()

    fig, ax = plt.subplots()
    bars = ax.bar(labels, counts, color=["green", "gray", "red"])
    ax.set_title("Sentiment Distribution")
    ax.set_xlabel("Sentiment")
    ax.set_ylabel("Count")

    # Add the count value at the top of each bar
    for bar in bars:
        yval = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            yval + 0.1,
            int(yval),
            ha="center",
            va="bottom",
        )

    # Save the graph as an image
    plt.savefig(os.path.join(output_folder, "sentiment_distribution.png"))
    plt.show()


# Main function to handle user input
if __name__ == "__main__":
    choice = (
        input(
            "Do you want to process a specific image or a folder? (Enter 'image' or 'folder'): "
        )
        .strip()
        .lower()
    )

    if choice == "folder":
        folder_path = input("Enter the folder path: ").strip()
        if os.path.isdir(folder_path):
            process_folder(folder_path)
        else:
            print(f"The folder '{folder_path}' does not exist.")

    elif choice == "image":
        image_path = input("Enter the image file path: ").strip()
        if os.path.isfile(image_path):
            print(f"Processing Image: {image_path}")

            # Extract text using EasyOCR
            extracted_text = extract_text_with_easyocr(image_path)
            print(f"Extracted Text: {extracted_text}")

            # Translate extracted text if it's in Tagalog (Filipino)
            translated_text = translate_text(
                extracted_text, "en"
            )  # Translate to English
            print(f"Translated Text: {translated_text}")

            # Analyze sentiment of the translated text
            (
                sentiment_translated,
                neg_translated,
                neu_translated,
                pos_translated,
                score_translated,
            ) = analyze_sentiment(translated_text)
            print(
                f"Sentiment of Translated Text: {sentiment_translated} (neg={neg_translated}, neu={neu_translated}, pos={pos_translated}, Score: {score_translated})"
            )

            # Generate a caption describing the image using BLIP
            image_caption = generate_caption(image_path)
            print(
                f"Image Caption: {image_caption}"
            )  # Print the generated description of the image

            # Analyze sentiment of the image caption
            sentiment_caption, neg_caption, neu_caption, pos_caption, score_caption = (
                analyze_sentiment(image_caption)
            )
            print(
                f"Sentiment of Image Caption: {sentiment_caption} (neg={neg_caption}, neu={neu_caption}, pos={pos_caption}, Score: {score_caption})"
            )

            # Calculate overall sentiment based on both scores (translated text + image caption)
            overall_score = score_translated + score_caption
            if overall_score >= 0.05:
                overall_sentiment = "Positive"
            elif overall_score <= -0.05:
                overall_sentiment = "Negative"
            else:
                overall_sentiment = "Neutral"

            print(f"Overall Sentiment: {overall_sentiment} (Score: {overall_score})")

            # Calculate confidence level as percentage
            confidence_level = calculate_confidence(overall_score)
            print(f"Confidence Level: {confidence_level}%")

        else:
            print(f"The file '{image_path}' does not exist.")

    else:
        print("Invalid choice! Please enter 'image' or 'folder'.")
