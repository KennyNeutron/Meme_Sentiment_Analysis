import os
import easyocr
from googletrans import Translator
from transformers import BlipProcessor, BlipForConditionalGeneration
from PIL import Image
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import torch
from torch.cuda.amp import autocast
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

# Initialize the VADER sentiment analyzer
analyzer = SentimentIntensityAnalyzer()


def extract_text_with_easyocr(image_path):
    """Extract text from an image using EasyOCR"""
    reader = easyocr.Reader(["en", "tl"], gpu=True)
    result = reader.readtext(image_path)
    extracted_text = " ".join([item[1] for item in result])
    return extracted_text


def translate_text(text, target_language="en"):
    """Translates extracted text to the target language"""
    if not text.strip():
        return ""
    try:
        translator = Translator()
        translated = translator.translate(text, dest=target_language)
        return translated.text
    except Exception as e:
        print(f"Error during translation: {e}")
        return text


def generate_caption(image_path):
    """Generate a caption using BLIP"""
    processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base")
    model = BlipForConditionalGeneration.from_pretrained(
        "Salesforce/blip-image-captioning-base"
    ).to("cuda")

    image = Image.open(image_path)
    inputs = processor(images=image, return_tensors="pt").to("cuda")
    with autocast():
        out = model.generate(**inputs)
    caption = processor.decode(out[0], skip_special_tokens=True)
    return caption


def analyze_sentiment(text):
    """Analyze sentiment using VADER"""
    sentiment = analyzer.polarity_scores(text)
    compound_score = sentiment["compound"]

    if compound_score >= 0.05:
        sentiment_label = "Positive"
    elif compound_score <= -0.05:
        sentiment_label = "Negative"
    else:
        sentiment_label = "Neutral"

    return (
        sentiment_label,
        sentiment["neg"],
        sentiment["neu"],
        sentiment["pos"],
        compound_score,
    )


def calculate_confidence(overall_score):
    """Calculate confidence level as percentage"""
    confidence = abs(overall_score) * 100
    if confidence < 50:
        confidence = 50 + (confidence * 0.5)
    else:
        confidence = 50 + (confidence**1.8) / 20
    confidence = min(confidence, 100)
    return round(confidence, 2)


def create_or_load_excel(output_folder, results):
    """Create or load the Excel file and append new results"""
    excel_file = os.path.join(output_folder, "sentiment_analysis_results.xlsx")

    # Sort results by file name
    results = results.sort_values(by="File Name")

    while True:
        try:
            if os.path.exists(excel_file):
                workbook = load_workbook(excel_file)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Sentiment Analysis Results"
                sheet.append(results.columns.tolist())

            for row in dataframe_to_rows(results, index=False, header=False):
                sheet.append(row)

            # Autofit column widths and align cells
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get column letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            workbook.save(excel_file)
            print(f"Results saved to {excel_file}")
            break
        except PermissionError:
            input(
                f"Please close the file '{excel_file}' and press Enter to continue..."
            )


def process_image(file_path):
    """Process a single image and return results"""
    try:
        print(f"Currently processing: {os.path.basename(file_path)}")
        extracted_text = extract_text_with_easyocr(file_path)

        if not extracted_text:
            translated_text = ""
            sentiment_translated = "Neutral"
            score_translated = 0
        else:
            translated_text = translate_text(extracted_text, "en")
            (
                sentiment_translated,
                neg_translated,
                neu_translated,
                pos_translated,
                score_translated,
            ) = analyze_sentiment(translated_text)

        image_caption = generate_caption(file_path)
        (
            sentiment_caption,
            neg_caption,
            neu_caption,
            pos_caption,
            score_caption,
        ) = analyze_sentiment(image_caption)

        overall_score = score_translated + score_caption
        overall_sentiment = (
            "Positive"
            if overall_score >= 0.05
            else "Negative" if overall_score <= -0.05 else "Neutral"
        )
        confidence_level = calculate_confidence(overall_score)

        return {
            "File Name": os.path.basename(file_path),
            "Extracted Text": extracted_text,
            "Translated Text": translated_text,
            "Text Sentiment": sentiment_translated,
            "Image Caption": image_caption,
            "Caption Sentiment": sentiment_caption,
            "Overall Sentiment": overall_sentiment,
            "Confidence": f"{confidence_level}%",
        }
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return None


def process_folder_parallel(folder_path):
    """Process images in a folder using parallel processing"""
    image_files = [
        os.path.join(folder_path, f)
        for f in os.listdir(folder_path)
        if f.lower().endswith((".jpg", ".jpeg", ".png"))
    ]

    results = []
    error_count = 0

    with ThreadPoolExecutor() as executor:
        futures = {executor.submit(process_image, file): file for file in image_files}
        for future in tqdm(
            as_completed(futures), total=len(image_files), desc="Processing Images"
        ):
            result = future.result()
            if result:
                results.append(result)
            else:
                error_count += 1

    results_df = pd.DataFrame(results)
    output_folder = os.path.dirname(folder_path)
    create_or_load_excel(output_folder, results_df)

    print(f"Processing complete. {error_count} image(s) were skipped due to errors.")


if __name__ == "__main__":
    choice = (
        input(
            "Do you want to process a specific image or a folder? (Enter 'image' or 'folder'): "
        )
        .strip()
        .lower()
    )

    if choice == "folder":
        folder_path = input("Enter the folder path: ").strip()
        if os.path.isdir(folder_path):
            process_folder_parallel(folder_path)
        else:
            print(f"The folder '{folder_path}' does not exist.")

    elif choice == "image":
        image_path = input("Enter the image file path: ").strip()
        if os.path.isfile(image_path):
            print(f"Processing Image: {image_path}")

            # Extract text using EasyOCR
            extracted_text = extract_text_with_easyocr(image_path)
            print(f"Extracted Text: {extracted_text}")

            # Translate extracted text if it's in Tagalog (Filipino)
            translated_text = translate_text(
                extracted_text, "en"
            )  # Translate to English
            print(f"Translated Text: {translated_text}")

            # Analyze sentiment of the translated text
            (
                sentiment_translated,
                neg_translated,
                neu_translated,
                pos_translated,
                score_translated,
            ) = analyze_sentiment(translated_text)
            print(
                f"Sentiment of Translated Text: {sentiment_translated} (neg={neg_translated}, neu={neu_translated}, pos={pos_translated}, Score: {score_translated})"
            )

            # Generate a caption describing the image using BLIP
            image_caption = generate_caption(image_path)
            print(
                f"Image Caption: {image_caption}"
            )  # Print the generated description of the image

            # Analyze sentiment of the image caption
            sentiment_caption, neg_caption, neu_caption, pos_caption, score_caption = (
                analyze_sentiment(image_caption)
            )
            print(
                f"Sentiment of Image Caption: {sentiment_caption} (neg={neg_caption}, neu={neu_caption}, pos={pos_caption}, Score: {score_caption})"
            )

            # Calculate overall sentiment based on both scores (translated text + image caption)
            overall_score = score_translated + score_caption
            if overall_score >= 0.05:
                overall_sentiment = "Positive"
            elif overall_score <= -0.05:
                overall_sentiment = "Negative"
            else:
                overall_sentiment = "Neutral"

            print(f"Overall Sentiment: {overall_sentiment} (Score: {overall_score})")

            # Calculate confidence level as percentage
            confidence_level = calculate_confidence(overall_score)
            print(f"Confidence Level: {confidence_level}%")

        else:
            print(f"The file '{image_path}' does not exist.")

    else:
        print("Invalid choice! Please enter 'image' or 'folder'.")
